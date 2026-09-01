import io
from pathlib import Path
from typing import Optional

from app.models import FileType


def parse_file(file_path: str, file_type: FileType) -> str:
    ext = file_type.value
    if ext == "pdf":
        return _parse_pdf(file_path)
    elif ext == "docx":
        return _parse_docx(file_path)
    elif ext == "xlsx":
        return _parse_xlsx(file_path)
    elif ext == "txt":
        return _parse_txt(file_path)
    elif ext == "md":
        return _parse_txt(file_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def _parse_pdf(file_path: str) -> str:
    import fitz
    doc = fitz.open(file_path)
    texts = []
    for page in doc:
        texts.append(page.get_text())
    doc.close()
    return "\n".join(texts)


def _parse_docx(file_path: str) -> str:
    from docx import Document
    doc = Document(file_path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs)


def _parse_xlsx(file_path: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    texts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        texts.append(f"[Sheet: {sheet}]")
        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
            if row_text.strip():
                texts.append(row_text)
    wb.close()
    return "\n".join(texts)


def _parse_txt(file_path: str) -> str:
    encodings = ["utf-8", "gbk", "gb2312", "latin-1"]
    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc) as f:
                return f.read()
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise ValueError("Unable to decode file with supported encodings")


def detect_file_type(filename: str) -> Optional[FileType]:
    ext = Path(filename).suffix.lower()
    mapping = {
        ".pdf": FileType.PDF,
        ".docx": FileType.DOCX,
        ".doc": FileType.DOCX,
        ".xlsx": FileType.XLSX,
        ".xls": FileType.XLSX,
        ".txt": FileType.TXT,
        ".md": FileType.MD,
    }
    return mapping.get(ext)
